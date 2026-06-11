"""
D4 tests: Excel parser — all edge cases.

Tests are self-contained: each creates an in-memory Excel file using openpyxl,
writes it to a temp file, and passes it to parse_salary_sheet().
No database, no Django views. The parser is pure logic.

Notes on openpyxl/pandas behaviour:
  - Cell formulas (=SUM()) are stored as NaN when read with dtype=str → treated as zero/blank
  - String-prefix injection (+50000, @cmd) IS detectable because pandas preserves the raw string
  - Negative numbers come through as '-5000' strings and must NOT be blocked
"""

import os
import tempfile
from decimal import Decimal

import openpyxl
from django.test import TestCase

from payroll.services.excel_parser import parse_salary_sheet


def _write_excel(rows: list[list], tmp_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(tmp_path)
    return tmp_path


def _parse(rows, emp_id_label="Employee", fixed=None, known=None):
    known = known or set()
    fixed = fixed or []
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        _write_excel(rows, path)
        return parse_salary_sheet(path, emp_id_label, fixed, known)
    finally:
        os.unlink(path)


class TestParserHappyPath(TestCase):
    def test_basic_two_employee_sheet(self):
        rows = [
            ["Employee", "EMP001", "EMP002"],
            ["Basic",    50000,    45000],
            ["HRA",      5000,     4000],
        ]
        result = _parse(rows, known={"EMP001", "EMP002"})
        self.assertFalse(result.has_fatal_errors)
        self.assertEqual(len(result.records), 2)
        emp1 = next(r for r in result.records if r.employee_number == "EMP001")
        self.assertEqual(emp1.breakdown["Basic"], Decimal("50000"))
        self.assertEqual(emp1.gross_total, Decimal("55000"))

    def test_first_row_employee_numbers_without_parser_config(self):
        rows = [
            ["", "EMP001", "EMP002"],
            ["Basic", 50000, 45000],
            ["Allowance", 5000, 4000],
        ]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            _write_excel(rows, path)
            result = parse_salary_sheet(path, known_employee_numbers={"EMP001", "EMP002"})
        finally:
            os.unlink(path)

        self.assertFalse(result.has_fatal_errors)
        self.assertEqual(len(result.records), 2)
        emp1 = next(r for r in result.records if r.employee_number == "EMP001")
        self.assertEqual(emp1.breakdown["Basic"], Decimal("50000"))

    def test_fixed_info_rows_are_skipped(self):
        rows = [
            ["Employee",    "EMP001"],
            ["Name",        "Nimal"],
            ["Designation", "Engineer"],
            ["Basic",       60000],
            ["HRA",         8000],
        ]
        result = _parse(rows, known={"EMP001"}, fixed=["Name", "Designation"])
        self.assertFalse(result.has_fatal_errors)
        rec = result.records[0]
        self.assertNotIn("Name", rec.breakdown)
        self.assertNotIn("Designation", rec.breakdown)
        self.assertIn("Basic", rec.breakdown)
        self.assertEqual(rec.gross_total, Decimal("68000"))

    def test_variable_component_count(self):
        rows = [
            ["Employee", "EMP001"],
            ["Basic",    40000],
            ["Travel",   3000],
            ["Medical",  2000],
        ]
        result = _parse(rows, known={"EMP001"})
        self.assertEqual(len(result.records[0].breakdown), 3)
        self.assertEqual(result.records[0].gross_total, Decimal("45000"))

    def test_different_id_row_label(self):
        rows = [
            ["Particulars", "EMP010"],
            ["Basic",        30000],
        ]
        result = _parse(rows, emp_id_label="Particulars", known={"EMP010"})
        self.assertFalse(result.has_fatal_errors)
        self.assertEqual(len(result.records), 1)


class TestParserEdgeCases(TestCase):
    def test_employee_number_with_whitespace(self):
        rows = [["Employee", "  EMP001  "], ["Basic", 50000]]
        result = _parse(rows, known={"EMP001"})
        self.assertEqual(result.records[0].employee_number, "EMP001")

    def test_employee_number_as_float(self):
        rows = [["Employee", 1.0], ["Basic", 50000]]
        result = _parse(rows, known={"1"})
        self.assertEqual(result.records[0].employee_number, "1")

    def test_blank_amount_treated_as_zero(self):
        rows = [["Employee", "EMP001"], ["Basic", 50000], ["Bonus", None]]
        result = _parse(rows, known={"EMP001"})
        self.assertEqual(result.records[0].breakdown["Bonus"], Decimal("0"))

    def test_amount_with_comma_separator(self):
        rows = [["Employee", "EMP001"], ["Basic", "50,000"]]
        result = _parse(rows, known={"EMP001"})
        self.assertEqual(result.records[0].breakdown["Basic"], Decimal("50000"))

    def test_formula_injection_plus_prefix_is_blocked(self):
        """String '+50000' written as text into an Excel cell should be blocked."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Employee"; ws["B1"] = "EMP001"
        ws["A2"] = "Basic"
        # Write as a plain string so pandas sees '+50000', not a formula result
        ws["B2"].value = "+50000"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        wb.save(path)
        try:
            result = parse_salary_sheet(path, "Employee", [], {"EMP001"})
        finally:
            os.unlink(path)
        self.assertEqual(result.records[0].breakdown["Basic"], Decimal("0"))
        self.assertTrue(any("formula injection" in w.lower() for w in result.warnings))

    def test_unparseable_amount_treated_as_zero_with_warning(self):
        """'N/A' written as plain text should warn and use zero."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Employee"; ws["B1"] = "EMP001"
        ws["A2"] = "Basic"; ws["B2"] = "N/A"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        wb.save(path)
        try:
            result = parse_salary_sheet(path, "Employee", [], {"EMP001"})
        finally:
            os.unlink(path)
        self.assertEqual(result.records[0].breakdown["Basic"], Decimal("0"))
        self.assertTrue(any("could not parse" in w.lower() for w in result.warnings))

    def test_unknown_employee_number_skipped_with_warning(self):
        rows = [["Employee", "EMP999"], ["Basic", 50000]]
        result = _parse(rows, known={"EMP001"})
        self.assertEqual(len(result.records), 0)
        self.assertTrue(any("EMP999" in w for w in result.warnings))

    def test_duplicate_employee_column_last_wins(self):
        rows = [
            ["Employee", "EMP001", "EMP001"],
            ["Basic",    30000,    50000],
        ]
        result = _parse(rows, known={"EMP001"})
        self.assertEqual(len(result.records), 1)
        self.assertEqual(result.records[0].breakdown["Basic"], Decimal("50000"))
        self.assertTrue(any("more than once" in w for w in result.warnings))

    def test_id_row_label_not_found_is_fatal(self):
        rows = [["WrongLabel", "EMP001"], ["Basic", 50000]]
        result = _parse(rows, emp_id_label="Employee", known={"EMP001"})
        self.assertTrue(result.has_fatal_errors)
        self.assertEqual(len(result.records), 0)

    def test_empty_file_is_fatal(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb = openpyxl.Workbook(); wb.active.title = "Empty"; wb.save(path)
            result = parse_salary_sheet(path, "Employee", [], set())
            self.assertTrue(result.has_fatal_errors)
        finally:
            os.unlink(path)

    def test_id_row_label_case_insensitive(self):
        rows = [["EMPLOYEE", "EMP001"], ["Basic", 50000]]
        result = _parse(rows, emp_id_label="Employee", known={"EMP001"})
        self.assertFalse(result.has_fatal_errors)
        self.assertEqual(len(result.records), 1)

    def test_fixed_labels_case_insensitive(self):
        rows = [
            ["Employee",      "EMP001"],
            ["EMPLOYEE NAME", "Nimal"],
            ["Basic",         50000],
        ]
        result = _parse(rows, known={"EMP001"}, fixed=["Employee Name"])
        self.assertNotIn("EMPLOYEE NAME", result.records[0].breakdown)
        self.assertIn("Basic", result.records[0].breakdown)

    def test_negative_amount_allowed(self):
        """Deductions are negative. The '-' prefix must NOT be blocked."""
        rows = [
            ["Employee",  "EMP001"],
            ["Basic",     50000],
            ["Deduction", -5000],
        ]
        result = _parse(rows, known={"EMP001"})
        self.assertEqual(result.records[0].breakdown["Deduction"], Decimal("-5000"))
        self.assertEqual(result.records[0].gross_total, Decimal("45000"))
